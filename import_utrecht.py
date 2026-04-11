"""
Import Natyanjani Utrecht (1).xlsx into BookKeeper.
Replaces all existing data with the latest information.

Client classification from 2026 sheet layout:
  Section 1 (rows 2-17):  Adults  - Wednesday offline group
  Section 2 (rows 22-34): Kids    - Wednesday offline group
  Section 3 (rows 39-42): Individual / private lessons
  Section 4 (rows 48-51): Adult tracking (Theertha/Shrujana/Dafodia/Olga)
    + Tuesday Online (Anju, Paalika, Nethra) from 2025 sheet
"""
import os
import sys
from datetime import date

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))
from app import create_app
from models import db, Client, Income, Expense, HourEntry

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Natyanjani Utrecht (1).xlsx")

# --------------- Name normalization ---------------
# Map every spelling variant → canonical display name
NAME_MAP = {
    "namith": "Namith",
    "kadjel": "Kadjel",
    "megha": "Megha",
    "urjitha": "Urjitha",
    "ashmitha": "Ashmitha",
    "ashmith": "Ashmitha",
    "anupama": "Anupama",
    "divya": "Divya",
    "sridevi": "Sridevi",
    "christy": "Christy",
    "chitra": "Chitra",
    "ambili": "Ambili",
    "lakshmi": "Lakshmi",
    "reethu": "Reethu",
    "kavana": "Kavana",
    "shivaranjani": "Shivaranjani",
    "vyasavi": "Vyasavi",
    "olga": "Olga",
    "theertha": "Theertha",
    "shrujana": "Shrujana",
    "aarna": "Aarna",
    "mihira": "Mihira",
    "anushka": "Anushka",
    "navya": "Navya",
    "nimar": "Nimar",
    "emlyn": "Emlyn",
    "dhanvi": "Dhanvi",
    "swara": "Swara",
    "anika": "Anika",
    "aaryahi": "Aaryahi",
    "aryahi": "Aaryahi",
    "gopika": "Gopika",
    "aadvika": "Aadvika",
    "advika": "Aadvika",
    "lagnajitha": "Lagnajitha",
    "shruthy": "Shruthy",
    "shruthy/offline": "Shruthy",
    "kavyah": "Kavyah",
    "kavyah/online": "Kavyah",
    "kaviah": "Kavyah",
    "kavyaah": "Kavyah",
    "sparsha": "Sparsha",
    "ankitha": "Ankitha",
    "ankitha/offline": "Ankitha",
    "shubh": "Shubh",
    "arjitha": "Arjitha",
    "manjusha": "Manjusha",
    "manjusha/online": "Manjusha",
    "dafodiya": "Dafodia",
    "dafodia": "Dafodia",
    "dafodya": "Dafodia",
    "anju": "Anju",
    "paalika": "Paalika",
    "palika": "Paalika",
    "nethra": "Nethra",
    "netra": "Nethra",
    "varshini": "Varshini",
    "leela": "Leela",
    "deepa": "Deepa",
    "amrutha": "Amrutha",
    "shrujana": "Shrujana",
    "manju": "Manjusha",
    "anjali": "Anjali",
}

# --------------- Client master list ---------------
# (canonical_name, student_type, status, notes)
# student_type: adult | child | individual | online
# status:       active | pending | discontinued
CLIENTS = [
    # --- Adults: Wednesday offline (2026 section 1) ---
    ("Namith",        "adult",      "active",       "Intensive group - Wednesday"),
    ("Kadjel",        "adult",      "active",       "Intensive group - Wednesday"),
    ("Megha",         "adult",      "active",       "Intensive group - Wednesday"),
    ("Urjitha",       "adult",      "active",       "Intensive group - Wednesday"),
    ("Ashmitha",      "adult",      "active",       "Intensive group - Wednesday"),
    ("Anupama",       "adult",      "active",       "Intensive group - Wednesday"),
    ("Divya",         "adult",      "active",       "Intensive group - Wednesday"),
    ("Sridevi",       "adult",      "active",       "Intensive group - Wednesday"),
    ("Christy",       "adult",      "active",       "Intensive group - Wednesday"),
    ("Chitra",        "adult",      "active",       "Intensive group - Wednesday; started Feb 2026"),
    ("Ambili",        "adult",      "active",       "Intensive group - Wednesday; started Feb 2026"),
    ("Lakshmi",       "adult",      "pending",      "Intensive group - Wednesday; registered, fees pending"),
    ("Reethu",        "adult",      "pending",      "Intensive group - Wednesday; registered, fees pending"),
    ("Kavana",        "adult",      "pending",      "Intensive group - Wednesday; registered, fees pending"),
    ("Shivaranjani",  "adult",      "pending",      "Intensive group - Wednesday; registered, fees pending"),
    ("Vyasavi",       "adult",      "pending",      "Intensive group - Wednesday; registered, fees pending"),
    ("Olga",          "adult",      "active",       "Intensive group - Wednesday"),
    ("Theertha",      "adult",      "active",       "Intensive group - Wednesday"),
    ("Shrujana",      "adult",      "active",       "Intensive group - Wednesday"),
    # --- Kids: Wednesday offline (2026 section 2) ---
    ("Aarna",         "child",      "active",       "Kids group - Wednesday"),
    ("Mihira",        "child",      "active",       "Kids group - Wednesday"),
    ("Anushka",       "child",      "active",       "Kids group - Wednesday"),
    ("Navya",         "child",      "active",       "Kids group - Wednesday"),
    ("Nimar",         "child",      "active",       "Kids group - Wednesday"),
    ("Emlyn",         "child",      "active",       "Kids group - Wednesday"),
    ("Dhanvi",        "child",      "active",       "Kids group - Wednesday; started Mar 2026"),
    ("Anika",         "child",      "active",       "Kids group - Wednesday"),
    ("Aaryahi",       "child",      "active",       "Kids group - Wednesday"),
    ("Gopika",        "child",      "active",       "Kids group - Wednesday"),
    ("Aadvika",       "child",      "active",       "Kids group - Wednesday"),
    ("Lagnajitha",    "child",      "pending",      "Kids group - Wednesday; registered, fees pending"),
    ("Swara",         "child",      "pending",      "Kids group - Wednesday; registered, fees pending"),
    # --- Individual / private (2026 section 3) ---
    ("Shruthy",       "individual", "active",       "Private offline lessons"),
    ("Kavyah",        "individual", "active",       "Private online lessons"),
    ("Sparsha",       "individual", "active",       "Private lessons"),
    ("Ankitha",       "individual", "active",       "Private offline lessons"),
    ("Manjusha",      "individual", "discontinued", "Private online; last active Aug 2025"),
    ("Shubh",         "individual", "discontinued", "Private lessons; last active Jul 2025"),
    ("Arjitha",       "individual", "discontinued", "Private lessons; last active Jun 2025"),
    # --- Online: Tuesday group (from 2025 sheet) ---
    ("Dafodia",       "online",     "active",       "Tuesday online group"),
    ("Anju",          "online",     "active",       "Tuesday online group"),
    ("Paalika",       "online",     "active",       "Tuesday online group"),
    ("Nethra",        "online",     "active",       "Tuesday online group"),
    # --- Discontinued ---
    ("Varshini",      "adult",      "discontinued", "Discontinued 2025"),
    ("Leela",         "adult",      "discontinued", "Discontinued 2025"),
    ("Deepa",         "online",     "discontinued", "Online Tuesday; discontinued 2025"),
    ("Amrutha",       "online",     "discontinued", "Online Tuesday; discontinued 2025"),
]


def canonical(raw):
    """Normalize a raw name from the sheet to canonical form."""
    if raw is None:
        return None
    return NAME_MAP.get(str(raw).strip().lower())


def val(row, col_idx):
    """Return a positive numeric value from a row at col_idx, or None."""
    if col_idx >= len(row):
        return None
    v = row[col_idx]
    if isinstance(v, (int, float)) and v > 0 and v < 1000:  # cap at 1000 to exclude totals/IBANs
        return round(float(v), 2)
    return None


def add_income(client_map, raw_name, year, month, amount, category, inv_prefix):
    cname = canonical(raw_name)
    if cname is None or amount is None:
        return 0
    client = client_map.get(cname)
    if client is None:
        return 0
    income = Income(
        client_id=client.id,
        date=date(year, month, 1),
        description=f"{category} ({_month_name(month)} {year})",
        amount=amount,
        btw_rate=0.0,
        btw_amount=0.0,
        total=amount,
        category=category,
        invoice_number=f"{inv_prefix}-{year}{month:02d}",
    )
    db.session.add(income)
    return 1


def _month_name(m):
    return ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][m - 1]


def import_2025_sheet(ws, client_map):
    """Parse the 2025 sheet: Wednesday offline, Individual, Online sections."""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    # ---- Section 1: Wednesday Offline (rows 6-30, 0-indexed 5-29) ----
    # Header row 5 (0-indexed 4): D=Feb, E=Mar ...N=Dec
    WED_COLS = {3:(2025,2),4:(2025,3),5:(2025,4),6:(2025,5),7:(2025,6),
                8:(2025,7),9:(2025,8),10:(2025,9),11:(2025,10),12:(2025,11),13:(2025,12)}
    for ri in range(5, 30):
        row = rows[ri]
        name = row[1]
        if not name:
            continue
        sln = row[0]
        cat = "Dance Classes - Intensive" if ri < 27 else "Dance Classes"
        for ci, (yr, mo) in WED_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci), cat,
                                f"W25-{str(name).strip()[:4].upper()}")

    # ---- Section 2: Individual Class Students (rows 39-45, 0-indexed 38-44) ----
    INDV_COLS = {4:(2025,2),5:(2025,3),6:(2025,4),7:(2025,5),8:(2025,6),
                 9:(2025,7),10:(2025,8),11:(2025,9),12:(2025,10),13:(2025,11),14:(2025,12)}
    for ri in range(38, 46):
        row = rows[ri]
        sln = row[2]
        name = row[3]
        if not name or not isinstance(sln, (int, float)):
            continue
        for ci, (yr, mo) in INDV_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Individual",
                                f"I25-{str(name).strip()[:4].upper()}")

    # ---- Section 3: Tuesday Online (rows 55-61, 0-indexed 54-60) ----
    ONL_COLS = {2:(2025,2),3:(2025,3),4:(2025,4),5:(2025,5),6:(2025,6),
                7:(2025,7),8:(2025,8),9:(2025,9),10:(2025,10),11:(2025,11),12:(2025,12)}
    for ri in range(54, 62):
        row = rows[ri]
        sln = row[0]
        name = row[1]
        if not name:
            continue
        for ci, (yr, mo) in ONL_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Online",
                                f"O25-{str(name).strip()[:4].upper()}")
    return count


def import_2026_sheet(ws, client_map):
    """Parse the 2026 sheet: all 4 sections."""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    # C=Jan2026 (col 2), D=Feb (col 3), E=Mar (col 4), ...
    Y26_COLS = {2:(2026,1),3:(2026,2),4:(2026,3),5:(2026,4),
                6:(2026,5),7:(2026,6),8:(2026,7),9:(2026,8),
                10:(2026,9),11:(2026,10),12:(2026,11)}

    # Section 1: Adults rows 2-17 (0-indexed 1-16)
    for ri in range(1, 17):
        row = rows[ri]
        name = row[1]
        if not name:
            continue
        for ci, (yr, mo) in Y26_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Intensive",
                                f"W26-{str(name).strip()[:4].upper()}")

    # Section 2: Kids rows 22-34 (0-indexed 21-33)
    for ri in range(21, 34):
        row = rows[ri]
        name = row[1]
        if not name:
            continue
        for ci, (yr, mo) in Y26_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Kids",
                                f"K26-{str(name).strip()[:4].upper()}")

    # Section 3: Individual rows 39-42 (0-indexed 38-41)
    for ri in range(38, 43):
        row = rows[ri]
        sln = row[0]
        name = row[1]
        if not name or not isinstance(sln, (int, float)):
            continue
        for ci, (yr, mo) in Y26_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Individual",
                                f"I26-{str(name).strip()[:4].upper()}")

    # Section 4: rows 48-51 (0-indexed 47-50)
    for ri in range(47, min(52, len(rows))):
        row = rows[ri]
        sln = row[0]
        name = row[1]
        if not name or not isinstance(sln, (int, float)):
            continue
        for ci, (yr, mo) in Y26_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Intensive",
                                f"W26-{str(name).strip()[:4].upper()}")

    return count


def import_2024_sheet(ws, client_map):
    """Parse limited 2024 data (Sep 2024 - Jan 2025)."""
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    # header row 3 (0-indexed 2): C=Sep2024, D=Oct, E=Nov, F=Dec, G=Jan2025
    Y24_COLS = {2:(2024,9),3:(2024,10),4:(2024,11),5:(2024,12),6:(2025,1)}
    for ri in range(3, 21):
        row = rows[ri]
        sln = row[0]
        name = row[1]
        if not name or not isinstance(sln, (int, float)):
            continue
        for ci, (yr, mo) in Y24_COLS.items():
            count += add_income(client_map, name, yr, mo, val(row, ci),
                                "Dance Classes - Intensive",
                                f"W24-{str(name).strip()[:4].upper()}")
    return count


def import_expenses_sheet(ws):
    """Import program expenses."""
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    # March 12th event — advance (col D=3) + balance (col E=4), rows 15-24 (0-indexed 14-23)
    for ri in range(14, 24):
        row = rows[ri]
        name = row[2]
        if not name:
            continue
        advance = row[3] if len(row) > 3 and isinstance(row[3], (int, float)) and row[3] > 0 else 0
        balance = row[4] if len(row) > 4 and isinstance(row[4], (int, float)) and row[4] > 0 else 0
        total = advance + balance
        if total > 0:
            db.session.add(Expense(
                date=date(2025, 3, 12),
                description=f"March 12th event - {str(name).strip()} (choreography & accessories)",
                amount=round(total, 2), btw_rate=0, btw_amount=0, total=round(total, 2),
                category="Events & Programs", receipt_ref=f"EVT-MAR12-{ri-13}",
            ))
            count += 1

    # Nov 9th Gajjee Pooje — May (col L=11) + Aug (col M=12), rows 11-31 (0-indexed 10-30)
    for ri in range(10, 31):
        row = rows[ri]
        sln = row[9]
        name = row[10]
        may = row[11] if len(row) > 11 and isinstance(row[11], (int, float)) and row[11] > 0 else 0
        aug_paid = row[12] if len(row) > 12 and isinstance(row[12], (int, float)) and row[12] > 0 else 0
        if not name or not isinstance(sln, (int, float)):
            continue
        name_s = str(name).strip()
        if may > 0:
            db.session.add(Expense(
                date=date(2025, 5, 1),
                description=f"Gajjee Pooje program - {name_s} (advance, May)",
                amount=may, btw_rate=0, btw_amount=0, total=may,
                category="Events & Programs", receipt_ref=f"EVT-GP-MAY-{sln:.0f}",
            ))
            count += 1
        if aug_paid > 0:
            db.session.add(Expense(
                date=date(2025, 8, 1),
                description=f"Gajjee Pooje program - {name_s} (balance, Aug)",
                amount=aug_paid, btw_rate=0, btw_amount=0, total=aug_paid,
                category="Events & Programs", receipt_ref=f"EVT-GP-AUG-{sln:.0f}",
            ))
            count += 1

    return count


def main():
    app = create_app()
    with app.app_context():
        print("Resetting database...")
        db.drop_all()
        db.create_all()

        # ---- Step 1: Create all clients ----
        print("\nCreating clients...")
        client_map = {}
        for name, stype, status, notes in CLIENTS:
            c = Client(name=name, student_type=stype, status=status, notes=notes)
            db.session.add(c)
            db.session.flush()
            client_map[name] = c
        db.session.flush()
        print(f"  {len(CLIENTS)} clients created")

        # ---- Step 2: Import income from all year sheets ----
        wb = load_workbook(EXCEL_FILE, data_only=True)

        print("\nImporting 2024 income...")
        cnt = import_2024_sheet(wb["2024"], client_map)
        print(f"  {cnt} income entries")

        print("Importing 2025 income...")
        cnt = import_2025_sheet(wb["2025"], client_map)
        print(f"  {cnt} income entries")

        print("Importing 2026 income...")
        cnt = import_2026_sheet(wb["2026"], client_map)
        print(f"  {cnt} income entries")

        # ---- Step 3: Import expenses ----
        print("\nImporting program expenses...")
        cnt = import_expenses_sheet(wb["program exp"])
        print(f"  {cnt} expense entries")

        wb.close()
        db.session.commit()

        # ---- Summary ----
        total_income = db.session.query(db.func.sum(Income.total)).scalar() or 0
        total_exp = db.session.query(db.func.sum(Expense.total)).scalar() or 0
        by_type = {}
        for c in Client.query.all():
            key = f"{c.student_type}/{c.status}"
            by_type[key] = by_type.get(key, 0) + 1

        print(f"\n{'='*55}")
        print("IMPORT COMPLETE")
        print(f"{'='*55}")
        print(f"Total clients:     {Client.query.count()}")
        for k, v in sorted(by_type.items()):
            print(f"  {k:<30s}: {v}")
        print(f"Total income:      EUR {total_income:,.2f}")
        print(f"Total expenses:    EUR {total_exp:,.2f}")
        print(f"{'='*55}")


if __name__ == "__main__":
    main()
