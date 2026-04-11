"""
Import Natyanjani dance fees Excel into BookKeeper database.
Reads student fees as Clients + Income, and program expenses as Expenses.
"""
import os
import sys
from datetime import date, datetime
from openpyxl import load_workbook

# Ensure we can import from the project
sys.path.insert(0, os.path.dirname(__file__))

from app import create_app
from models import db, Client, Income, Expense

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Summary of dance fees Natyanjani.xlsx")

# Month mapping for columns in the spreadsheet
MONTH_MAP_INTENSIVE = {
    # col_index: (year, month)
    2: (2025, 1),   # C = January
    3: (2025, 2),   # D = February
    4: (2025, 3),   # E = March
    5: (2025, 4),   # F = April
    6: (2025, 5),   # G = May
    7: (2025, 6),   # H = June
    8: (2025, 7),   # I = July
    9: (2025, 8),   # J = August
    10: (2025, 9),  # K = September
}

MONTH_MAP_PRIVATE = {
    # col_index: (year, month) — Private section starts from Feb (col C)
    2: (2025, 2),   # C = February
    3: (2025, 3),   # D = March
    4: (2025, 4),   # E = April
    5: (2025, 5),   # F = May
    6: (2025, 6),   # G = June
    7: (2025, 7),   # H = July
    8: (2025, 8),   # I = August
    9: (2025, 9),   # J = September
}

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def get_or_create_client(name, notes=""):
    """Find existing client by name or create new one."""
    name = name.strip()
    client = Client.query.filter(Client.name == name).first()
    if not client:
        client = Client(name=name, notes=notes)
        db.session.add(client)
        db.session.flush()  # Get the ID
    elif notes and not client.notes:
        client.notes = notes
    return client


def import_fees_sheet(wb):
    """Import fees collected 2025 sheet — creates clients and income records."""
    ws = wb["fees collected 2025"]
    rows = list(ws.iter_rows(values_only=True))

    clients_added = 0
    income_added = 0

    # --- Section 1: INTENSIVE (8 classes/month) — rows 5-19 (index 4-18) ---
    for row_idx in range(4, 19):  # rows 5 to 19
        row = rows[row_idx]
        sln = row[0]
        name = row[1]
        bank_info = row[14] if len(row) > 14 else None  # Column O

        if not name or not isinstance(sln, (int, float)):
            continue

        name = str(name).strip()
        notes = f"Intensive (8 classes/month) | Bank: {bank_info.strip()}" if bank_info else "Intensive (8 classes/month)"
        client = get_or_create_client(name, notes)
        clients_added += 1

        for col_idx, (year, month) in MONTH_MAP_INTENSIVE.items():
            amount = row[col_idx] if col_idx < len(row) else None
            if amount and isinstance(amount, (int, float)) and amount > 0:
                income = Income(
                    client_id=client.id,
                    date=date(year, month, 1),
                    description=f"Dance class fee - Intensive ({MONTH_NAMES[month]} {year})",
                    amount=round(amount, 2),
                    btw_rate=0.0,
                    btw_amount=0.0,
                    total=round(amount, 2),
                    category="Dance Classes - Intensive",
                    invoice_number=f"NAT-I-{sln:02.0f}-{year}{month:02d}",
                )
                db.session.add(income)
                income_added += 1

    # --- Section 2: BASIC COURSE (4 classes/month) — rows 23-26 (index 22-25) ---
    for row_idx in range(22, 26):
        row = rows[row_idx]
        sln = row[0]
        name = row[1]
        bank_info = row[14] if len(row) > 14 else None

        if not name or not isinstance(sln, (int, float)):
            continue

        name = str(name).strip()
        notes = f"Basic (4 classes/month) | Bank: {bank_info.strip()}" if bank_info else "Basic (4 classes/month)"
        client = get_or_create_client(name, notes)
        clients_added += 1

        for col_idx, (year, month) in MONTH_MAP_INTENSIVE.items():
            amount = row[col_idx] if col_idx < len(row) else None
            if amount and isinstance(amount, (int, float)) and amount > 0:
                income = Income(
                    client_id=client.id,
                    date=date(year, month, 1),
                    description=f"Dance class fee - Basic ({MONTH_NAMES[month]} {year})",
                    amount=round(amount, 2),
                    btw_rate=0.0,
                    btw_amount=0.0,
                    total=round(amount, 2),
                    category="Dance Classes - Basic",
                    invoice_number=f"NAT-B-{sln:02.0f}-{year}{month:02d}",
                )
                db.session.add(income)
                income_added += 1

    # --- Discontinued students — rows 28-29 (index 27-28) ---
    for row_idx in range(27, 29):
        row = rows[row_idx]
        name = row[1]
        if not name:
            continue
        name = str(name).strip()
        client = get_or_create_client(name, "Discontinued")

        for col_idx, (year, month) in MONTH_MAP_INTENSIVE.items():
            amount = row[col_idx] if col_idx < len(row) else None
            if amount and isinstance(amount, (int, float)) and amount > 0:
                income = Income(
                    client_id=client.id,
                    date=date(year, month, 1),
                    description=f"Dance class fee ({MONTH_NAMES[month]} {year})",
                    amount=round(amount, 2),
                    btw_rate=0.0,
                    btw_amount=0.0,
                    total=round(amount, 2),
                    category="Dance Classes",
                    invoice_number=f"NAT-D-{year}{month:02d}",
                )
                db.session.add(income)
                income_added += 1

    # --- Section 3: PRIVATE STUDENTS — rows 38-40 (index 37-39) ---
    for row_idx in range(37, 40):
        row = rows[row_idx]
        sln = row[0]
        name = row[1]
        # Bank info is in column N (index 13) for private
        bank_info = row[13] if len(row) > 13 else None

        if not name or not isinstance(sln, (int, float)):
            continue

        name = str(name).strip()
        notes = f"Private student | Bank: {bank_info.strip()}" if bank_info else "Private student"
        client = get_or_create_client(name, notes)
        clients_added += 1

        for col_idx, (year, month) in MONTH_MAP_PRIVATE.items():
            amount = row[col_idx] if col_idx < len(row) else None
            if amount and isinstance(amount, (int, float)) and amount > 0:
                income = Income(
                    client_id=client.id,
                    date=date(year, month, 1),
                    description=f"Private dance class fee ({MONTH_NAMES[month]} {year})",
                    amount=round(amount, 2),
                    btw_rate=0.0,
                    btw_amount=0.0,
                    total=round(amount, 2),
                    category="Dance Classes - Private",
                    invoice_number=f"NAT-P-{sln:02.0f}-{year}{month:02d}",
                )
                db.session.add(income)
                income_added += 1

    return clients_added, income_added


def import_expenses_sheet(wb):
    """Import program expenses sheet."""
    ws = wb["program exp"]
    rows = list(ws.iter_rows(values_only=True))

    expenses_added = 0

    # --- March 12th event — advance + balance columns (D=advance, E=balance) ---
    # Rows 15-24 (index 14-23 in 0-based, but rows list is 0-based from row 1)
    # Row 11 in sheet = index 10 in rows list... Let me re-check.
    # ws rows are 1-based, list is 0-based: rows[0] = row 1
    # Row 14 (header): index 13, Row 15: index 14, etc.

    for row_idx in range(14, 24):  # sheet rows 15-24
        row = rows[row_idx]
        name = row[2]  # Column C
        advance = row[3]  # Column D
        balance = row[4]  # Column E

        if not name:
            continue

        total = 0
        if advance and isinstance(advance, (int, float)):
            total += advance
        if balance and isinstance(balance, (int, float)):
            total += balance

        if total > 0:
            expense = Expense(
                date=date(2025, 3, 12),
                description=f"March 12th event - {str(name).strip()} (choreography classes & accessories)",
                amount=round(total, 2),
                btw_rate=0.0,
                btw_amount=0.0,
                total=round(total, 2),
                category="Events & Programs",
                receipt_ref=f"EVT-MAR12-{row_idx - 13}",
            )
            db.session.add(expense)
            expenses_added += 1

    # --- Nov 9th Gajjee Pooje program — May (J) and June (K) columns ---
    # Rows 11-30 (index 10-29), columns H-K
    for row_idx in range(10, 30):
        row = rows[row_idx]
        sln = row[7]   # Column H
        name = row[8]   # Column I
        may_amt = row[9] if len(row) > 9 else None   # Column J
        june_amt = row[10] if len(row) > 10 else None  # Column K

        if not name or not isinstance(sln, (int, float)):
            continue

        name = str(name).strip()

        if may_amt and isinstance(may_amt, (int, float)) and may_amt > 0:
            expense = Expense(
                date=date(2025, 5, 1),
                description=f"Gajjee Pooje program prep - {name} (May)",
                amount=round(may_amt, 2),
                btw_rate=0.0,
                btw_amount=0.0,
                total=round(may_amt, 2),
                category="Events & Programs",
                receipt_ref=f"EVT-GP-MAY-{sln:.0f}",
            )
            db.session.add(expense)
            expenses_added += 1

        if june_amt and isinstance(june_amt, (int, float)) and june_amt > 0:
            expense = Expense(
                date=date(2025, 6, 1),
                description=f"Gajjee Pooje program prep - {name} (June)",
                amount=round(june_amt, 2),
                btw_rate=0.0,
                btw_amount=0.0,
                total=round(june_amt, 2),
                category="Events & Programs",
                receipt_ref=f"EVT-GP-JUN-{sln:.0f}",
            )
            db.session.add(expense)
            expenses_added += 1

    return expenses_added


def main():
    app = create_app()
    with app.app_context():
        print(f"Reading Excel: {EXCEL_FILE}")
        wb = load_workbook(EXCEL_FILE, data_only=True)

        print("\n--- Importing fees (Clients + Income) ---")
        clients_added, income_added = import_fees_sheet(wb)

        print("\n--- Importing program expenses ---")
        expenses_added = import_expenses_sheet(wb)

        db.session.commit()
        wb.close()

        # Print summary
        total_clients = Client.query.count()
        total_income = db.session.query(db.func.sum(Income.total)).scalar() or 0
        total_expenses = db.session.query(db.func.sum(Expense.total)).scalar() or 0

        print(f"\n{'='*50}")
        print(f"IMPORT COMPLETE")
        print(f"{'='*50}")
        print(f"Clients in database:  {total_clients}")
        print(f"Income entries added: {income_added}")
        print(f"Expense entries added: {expenses_added}")
        print(f"Total income:   EUR {total_income:,.2f}")
        print(f"Total expenses: EUR {total_expenses:,.2f}")
        print(f"{'='*50}")


if __name__ == "__main__":
    main()
