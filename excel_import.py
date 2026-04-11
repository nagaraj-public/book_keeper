from openpyxl import load_workbook
from models import db, Client


# Common column name mappings (case-insensitive)
COLUMN_MAP = {
    "name": ["name", "naam", "client name", "klant", "contact"],
    "company_name": ["company", "company name", "bedrijfsnaam", "bedrijf", "firma"],
    "kvk_number": ["kvk", "kvk number", "kvk nummer", "kvk-nummer"],
    "btw_number": ["btw", "btw number", "btw nummer", "btw-nummer", "vat", "vat number"],
    "email": ["email", "e-mail", "emailadres", "e-mailadres"],
    "phone": ["phone", "telefoon", "tel", "telephone", "phone number", "telefoonnummer"],
    "address": ["address", "adres", "street", "straat"],
    "notes": ["notes", "opmerkingen", "notities", "remarks"],
}


def _match_column(header):
    """Match a header string to a Client field name."""
    header_lower = header.strip().lower()
    for field, aliases in COLUMN_MAP.items():
        if header_lower in aliases:
            return field
    return None


def import_clients_from_excel(filepath):
    """Import clients from an Excel file. Returns the number of clients imported."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty.")

    # Map headers
    headers = rows[0]
    field_map = {}
    for idx, header in enumerate(headers):
        if header is None:
            continue
        field = _match_column(str(header))
        if field:
            field_map[idx] = field

    if not field_map:
        raise ValueError(
            "Could not match any column headers. "
            "Expected columns like: Name, Company, KVK, BTW, Email, Phone, Address, Notes"
        )

    count = 0
    for row in rows[1:]:
        data = {}
        for idx, field in field_map.items():
            if idx < len(row) and row[idx] is not None:
                data[field] = str(row[idx]).strip()

        if not data.get("name"):
            # Try company_name as fallback
            if data.get("company_name"):
                data["name"] = data["company_name"]
            else:
                continue  # Skip rows without a name

        client = Client(**data)
        db.session.add(client)
        count += 1

    db.session.commit()
    wb.close()
    return count
